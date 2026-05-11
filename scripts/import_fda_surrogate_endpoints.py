"""Import FDA surrogate endpoint table workbook into dismech YAML."""

from __future__ import annotations

import argparse
import hashlib
import re
from datetime import date
from pathlib import Path
from typing import Any

import openpyxl
import yaml


FDA_SOURCE_URL = (
    "https://www.fda.gov/drugs/development-resources/"
    "table-surrogate-endpoints-were-basis-drug-approval-or-licensure"
)
FDA_WORKBOOK_URL = "https://www.fda.gov/media/156533/download?attachment="

TABLES = {
    "Adult SE Non-cancer Related": {
        "code": "ADULT_NONCANCER",
        "prefix": "adult-noncancer",
        "has_age_range": False,
    },
    "Adult SE Cancer Related": {
        "code": "ADULT_CANCER",
        "prefix": "adult-cancer",
        "has_age_range": False,
    },
    "Pediatric Non-cancer Related": {
        "code": "PEDIATRIC_NONCANCER",
        "prefix": "pediatric-noncancer",
        "has_age_range": True,
    },
    "Pediatric Cancer Related": {
        "code": "PEDIATRIC_CANCER",
        "prefix": "pediatric-cancer",
        "has_age_range": True,
    },
}

FOOTNOTE_MARKERS = {
    "#": "COMPOSITE_BIOMARKER_SURROGATE",
    "*": "MECHANISM_AGNOSTIC",
    "§": "TUMOR_BURDEN_CONTEXT_DEPENDENT",
    "˟": "ANTICIPATED_PRIMARY_EFFICACY_USE",
    "¤": "BONE_MINERAL_DENSITY_CONTEXT",
    "†": "CLINICAL_ENDPOINTS_REQUIRED",
}

MANUAL_MAPPINGS = {
    "duchenne muscular dystrophy dmd": (
        "CURATED_DISMECH_MAPPING",
        ["Duchenne Muscular Dystrophy"],
        ["kb/disorders/Duchenne_Muscular_Dystrophy.yaml"],
        "Curated match from FDA abbreviation-bearing label.",
    ),
    "fabry disease": (
        "EXACT_DISMECH_MATCH",
        ["Fabry disease"],
        ["kb/disorders/Fabry_Disease.yaml"],
        "Exact disease label match.",
    ),
    "type 1 diabetes mellitus": (
        "CURATED_DISMECH_MAPPING",
        ["Type I Diabetes"],
        ["kb/disorders/Type_I_Diabetes.yaml"],
        "Curated match from FDA Arabic-numeral label to dismech Roman-numeral disease name.",
    ),
}

PATIENT_POPULATION_CANDIDATES = [
    (
        re.compile(r"\bsickle cell disease\b", re.IGNORECASE),
        "Sickle Cell Disease",
        "kb/disorders/Sickle_Cell_Disease.yaml",
    ),
    (
        re.compile(r"\bthalassemia\b", re.IGNORECASE),
        "Beta Thalassemia",
        "kb/disorders/Beta_Thalassemia.yaml",
    ),
]

NOT_DISEASE_SPECIFIC_PATTERNS = [
    re.compile(pattern, re.IGNORECASE)
    for pattern in (
        r"\bvaccine\b",
        r"\bantiseptic\b",
        r"\bimmune globulin\b",
        r"\banticoagulation reversal\b",
        r"\btobacco dependence\b",
        r"\bopioid use disorder\b",
        r"\binteroperative hemorrhage\b",
        r"\bsupportive cancer care\b",
    )
]


def _norm(value: str) -> str:
    value = value.replace("\n", " ")
    value = re.sub(r"\([^)]*\)", " ", value)
    value = re.sub(r"[^A-Za-z0-9]+", " ", value)
    return re.sub(r"\s+", " ", value).strip().lower()


def _cell(value: Any) -> str:
    if value is None:
        return ""
    return re.sub(r"\s+", " ", str(value)).strip()


def _remove_footnote_markers(value: str) -> str:
    for marker in FOOTNOTE_MARKERS:
        value = value.replace(marker, "")
    value = value.replace("**", "")
    return re.sub(r"\s+", " ", value).strip()


def _footnotes(*values: str) -> list[str]:
    joined = " ".join(values)
    notes = {
        footnote for marker, footnote in FOOTNOTE_MARKERS.items() if marker in joined
    }
    if "**" in joined:
        notes.add("ARRHYTHMIA_RESPONSE_DEFINITION")
    return sorted(notes)


def _approval_type(raw_value: str) -> str:
    cleaned = _remove_footnote_markers(raw_value).lower()
    cleaned = cleaned.replace(" ", "")
    cleaned = cleaned.replace("trasi" + "tional", "traditional")
    if cleaned in {
        "accelerated/traditional",
        "traditional/accelerated",
        "acceleratedortraditional",
    }:
        return "ACCELERATED_OR_TRADITIONAL"
    if "monograph" in cleaned:
        return "TRADITIONAL_AND_MONOGRAPH"
    if cleaned == "accelerated":
        return "ACCELERATED"
    if cleaned == "traditional":
        return "TRADITIONAL"
    raise ValueError(f"Unexpected approval type: {raw_value!r}")


def _validation_level(approval_type: str) -> str:
    if approval_type == "ACCELERATED":
        return "REASONABLY_LIKELY_SURROGATE_ENDPOINT"
    if approval_type == "TRADITIONAL":
        return "VALIDATED_SURROGATE_ENDPOINT"
    return "CONTEXT_DEPENDENT_SURROGATE_ENDPOINT"


def _clinical_benefit_linkage(approval_type: str) -> str:
    if approval_type == "ACCELERATED":
        return "REASONABLY_LIKELY_TO_PREDICT_CLINICAL_BENEFIT"
    if approval_type == "TRADITIONAL":
        return "KNOWN_TO_PREDICT_CLINICAL_BENEFIT"
    return "CONTEXT_DEPENDENT"


def _clinical_benefit_basis(approval_type: str) -> str:
    if approval_type == "ACCELERATED":
        return (
            "FDA table lists this row as accelerated approval; under BEST/FDA "
            "semantics this indicates a surrogate endpoint reasonably likely to "
            "predict clinical benefit in the stated context of use."
        )
    if approval_type == "TRADITIONAL":
        return (
            "FDA table lists this row as traditional approval; under BEST/FDA "
            "semantics this indicates a surrogate endpoint known to predict "
            "clinical benefit in the stated context of use."
        )
    return (
        "FDA table lists this row as usable for accelerated and/or traditional "
        "approval depending on context of use; clinical-benefit linkage is "
        "therefore context dependent."
    )


def _load_disorder_index(kb_dir: Path) -> dict[str, tuple[str, str]]:
    index: dict[str, tuple[str, str]] = {}
    if not kb_dir.exists():
        return index
    for path in sorted(kb_dir.glob("*.yaml")):
        if path.name.endswith(".history.yaml"):
            continue
        try:
            data = yaml.safe_load(path.read_text())
        except Exception:
            continue
        if not isinstance(data, dict) or not data.get("name"):
            continue
        name = str(data["name"])
        index[_norm(name)] = (name, str(path))
    return index


def _mapping(
    disease_or_use: str,
    patient_population: str,
    disorder_index: dict[str, tuple[str, str]],
) -> dict[str, Any]:
    normalized = _norm(disease_or_use)
    if normalized in MANUAL_MAPPINGS:
        status, diseases, files, notes = MANUAL_MAPPINGS[normalized]
        return {
            "mapping_status": status,
            "mapped_diseases": diseases,
            "mapped_disease_files": files,
            "mapping_notes": notes,
        }
    if normalized in disorder_index:
        disease_name, disease_file = disorder_index[normalized]
        return {
            "mapping_status": "EXACT_DISMECH_MATCH",
            "mapped_diseases": [disease_name],
            "mapped_disease_files": [disease_file],
            "mapping_notes": "Exact normalized FDA disease/use label match.",
        }

    candidate_diseases: list[str] = []
    candidate_files: list[str] = []
    combined = f"{disease_or_use} {patient_population}"
    for pattern, disease_name, disease_file in PATIENT_POPULATION_CANDIDATES:
        if pattern.search(combined):
            candidate_diseases.append(disease_name)
            candidate_files.append(disease_file)
    if candidate_diseases:
        return {
            "mapping_status": "CANDIDATE_DISMECH_MAPPING",
            "mapped_diseases": candidate_diseases,
            "mapped_disease_files": candidate_files,
            "mapping_notes": (
                "Candidate mapping inferred from disease mention in the FDA "
                "disease/use or patient-population text; requires curator review."
            ),
        }

    if any(pattern.search(disease_or_use) for pattern in NOT_DISEASE_SPECIFIC_PATTERNS):
        return {
            "mapping_status": "NOT_DISEASE_SPECIFIC",
            "mapping_notes": (
                "FDA row is a product-use, prevention, or broad regulatory "
                "context rather than a directly mapped dismech disease entry."
            ),
        }

    return {
        "mapping_status": "NEEDS_CURATION",
        "mapping_notes": "No dismech disease mapping assigned during initial FDA import.",
    }


def _context_of_use(row: dict[str, str]) -> str:
    parts = [
        f"Disease/use: {row['disease_or_use']}",
        f"population: {row['patient_population']}",
        f"endpoint: {row['surrogate_endpoint']}",
        f"approval context: {row['approval_type'].replace('_', ' ').lower()}",
        f"drug mechanism: {row['drug_mechanism_of_action']}",
    ]
    if row.get("age_range"):
        parts.append(f"age range: {row['age_range']}")
    return "; ".join(parts) + "."


def import_workbook(
    workbook_path: Path,
    kb_dir: Path,
    content_current_as_of: str,
    retrieved_date: str,
) -> dict[str, Any]:
    workbook_bytes = workbook_path.read_bytes()
    workbook_sha256 = hashlib.sha256(workbook_bytes).hexdigest()
    disorder_index = _load_disorder_index(kb_dir)
    with workbook_path.open("rb") as handle:
        workbook = openpyxl.load_workbook(handle, data_only=True)

    surrogate_endpoints: list[dict[str, Any]] = []
    for sheet_name, config in TABLES.items():
        sheet = workbook[sheet_name]
        row_counter = 0
        for row_number in range(2, sheet.max_row + 1):
            disease_or_use = _cell(sheet.cell(row_number, 1).value)
            if not disease_or_use:
                continue
            patient_population = _cell(sheet.cell(row_number, 2).value)
            endpoint = _cell(sheet.cell(row_number, 3).value)
            approval_raw = _cell(sheet.cell(row_number, 4).value)
            mechanism = _cell(sheet.cell(row_number, 5).value)
            age_range = _cell(sheet.cell(row_number, 6).value)
            approval_type = _approval_type(approval_raw)
            row_counter += 1
            row_id = f"FDA-SE-{config['prefix']}-{row_counter:03d}"
            row = {
                "row_id": row_id,
                "source_table": config["code"],
                "source_sheet": sheet_name,
                "source_row_number": row_number,
                "disease_or_use": disease_or_use,
                "patient_population": patient_population,
                "surrogate_endpoint": endpoint,
                "approval_type": approval_type,
                "drug_mechanism_of_action": _remove_footnote_markers(mechanism),
                "endpoint_validation_level": _validation_level(approval_type),
                "clinical_benefit_linkage": _clinical_benefit_linkage(approval_type),
                "clinical_benefit_linkage_basis": _clinical_benefit_basis(
                    approval_type
                ),
                "source_url": FDA_SOURCE_URL,
                "source_workbook_url": FDA_WORKBOOK_URL,
                "source_workbook_sha256": workbook_sha256,
                "source_content_current_as_of": content_current_as_of,
                "retrieved_date": retrieved_date,
            }
            if config["has_age_range"] and age_range:
                row["age_range"] = age_range
            notes = _footnotes(endpoint, approval_raw, mechanism)
            if notes:
                row["footnotes"] = notes
            row["context_of_use"] = _context_of_use(row)
            row.update(_mapping(disease_or_use, patient_population, disorder_index))
            surrogate_endpoints.append(row)

    return {
        "name": "FDA Surrogate Endpoint Table",
        "description": (
            "Curated row-level import of FDA's Surrogate Endpoint Table for "
            "drug and biologic development. Rows preserve the FDA disease/use, "
            "patient-population, surrogate-endpoint, approval-pathway, drug "
            "mechanism, and pediatric age-range context."
        ),
        "source_url": FDA_SOURCE_URL,
        "source_workbook_url": FDA_WORKBOOK_URL,
        "source_workbook_sha256": workbook_sha256,
        "source_content_current_as_of": content_current_as_of,
        "retrieved_date": retrieved_date,
        "tracked_issues": [
            {
                "url": "https://github.com/monarch-initiative/dismech/issues/2495",
                "title": "Incorporate FDA surrogate endpoint table and BEST endpoint semantics",
                "tracked_issue_role": "curation_scope",
                "tracked_issue_status": "open",
            }
        ],
        "surrogate_endpoints": surrogate_endpoints,
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    parser.add_argument("--kb-dir", default=Path("kb/disorders"), type=Path)
    parser.add_argument("--content-current-as-of", default="2026-04-29")
    parser.add_argument("--retrieved-date", default=date.today().isoformat())
    args = parser.parse_args()

    collection = import_workbook(
        workbook_path=args.input,
        kb_dir=args.kb_dir,
        content_current_as_of=args.content_current_as_of,
        retrieved_date=args.retrieved_date,
    )
    args.output.parent.mkdir(parents=True, exist_ok=True)
    args.output.write_text(
        yaml.safe_dump(
            collection,
            sort_keys=False,
            allow_unicode=True,
            width=100,
        )
    )


if __name__ == "__main__":
    main()
